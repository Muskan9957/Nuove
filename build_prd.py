from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── colours ──
C_NAV   = "0D1B2A"   # deep navy
C_WHITE = "FFFFFF"
C_P1    = "C0392B"
C_P2    = "E67E22"
C_P3    = "27AE60"
C_P4    = "2980B9"
C_ALT   = "F4F6FB"
C_HEAD  = "1A2740"
C_SECT  = "DDE3F0"
C_DONE  = "EAFAF1"
C_NOTE  = "FEF9E7"

PMAP = {
    "P1 – Must Have":  (C_P1, C_WHITE),
    "P2 – Should Have":(C_P2, C_WHITE),
    "P3 – Nice to Have":(C_P3,C_WHITE),
    "P4 – Future":     (C_P4, C_WHITE),
}

def tb():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hf(sz=10, bold=True, color=C_WHITE):
    return Font(name="Arial", bold=bold, size=sz, color=color)

def bf(sz=9, bold=False, color="1A1A1A"):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def fill(h):
    return PatternFill("solid", fgColor=h)

def ca(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ════════════════════════════════════════════════════
# SHEET 1 — COVER
# ════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "Cover"
ws0.sheet_view.showGridLines = False

def mrow(ws, row, cols, val, bg, fg="FFFFFF", sz=12, bold=True, h="center"):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws[f"A{row}"]
    c.value = val
    c.font = Font(name="Arial", bold=bold, size=sz, color=fg)
    c.fill = fill(bg)
    c.alignment = ca(h)

mrow(ws0, 1, 8, "NUOVE — PRODUCT REQUIREMENTS DOCUMENT", C_NAV, sz=20, bold=True)
ws0.row_dimensions[1].height = 50
mrow(ws0, 2, 8, "Version 1.0  |  App Store & Play Store Launch  |  Prepared by: Product Owner", C_HEAD, sz=11)
ws0.row_dimensions[2].height = 28
mrow(ws0, 3, 8, "", C_NAV)
ws0.row_dimensions[3].height = 8

# About section
about = [
    ("App Name",        "Nuove"),
    ("Domain",          "nuove.in"),
    ("Platform",        "iOS App Store + Google Play Store + Web (nuove.in)"),
    ("Frontend",        "React + Vite → Vercel  |  Capacitor wrapper for native apps"),
    ("Backend",         "Node.js + Express → Railway"),
    ("Database",        "PostgreSQL (to be provisioned)"),
    ("Payments",        "Razorpay (India)"),
    ("Target Users",    "Indian short-form video creators (Instagram Reels, YouTube Shorts)"),
    ("Target Launch",   "TBD — all P1 items must be complete before submission"),
]
mrow(ws0, 4, 8, "APP OVERVIEW", C_HEAD, sz=11)
ws0.row_dimensions[4].height = 22
for i, (k, v) in enumerate(about, start=5):
    ws0.merge_cells(f"A{i}:C{i}")
    ws0[f"A{i}"] = k
    ws0[f"A{i}"].font = hf(10, bold=True, color="1A1A1A")
    ws0[f"A{i}"].fill = fill(C_SECT)
    ws0[f"A{i}"].alignment = ca("left")
    ws0[f"A{i}"].border = tb()
    ws0.merge_cells(f"D{i}:H{i}")
    ws0[f"D{i}"] = v
    ws0[f"D{i}"].font = bf(10)
    ws0[f"D{i}"].alignment = ca("left")
    ws0[f"D{i}"].border = tb()
    ws0.row_dimensions[i].height = 20

# Priority legend
lr = len(about) + 6
mrow(ws0, lr, 8, "PRIORITY GUIDE", C_HEAD, sz=11)
ws0.row_dimensions[lr].height = 22
legends = [
    ("P1 – Must Have",   C_P1, "Blocker — app CANNOT launch without this. Do first."),
    ("P2 – Should Have", C_P2, "Important — do within Week 1 of development sprint."),
    ("P3 – Nice to Have",C_P3, "Good UX improvement — do within Month 1."),
    ("P4 – Future",      C_P4, "Post-launch backlog. Schedule after first 500 users."),
]
for i, (lbl, bg, desc) in enumerate(legends, start=lr+1):
    ws0.merge_cells(f"A{i}:C{i}")
    ws0[f"A{i}"] = lbl
    ws0[f"A{i}"].font = hf(10)
    ws0[f"A{i}"].fill = fill(bg)
    ws0[f"A{i}"].alignment = ca()
    ws0[f"A{i}"].border = tb()
    ws0.merge_cells(f"D{i}:H{i}")
    ws0[f"D{i}"] = desc
    ws0[f"D{i}"].font = bf(10)
    ws0[f"D{i}"].alignment = ca("left")
    ws0[f"D{i}"].border = tb()
    ws0.row_dimensions[i].height = 20

# Module summary
mr = lr + len(legends) + 2
mrow(ws0, mr, 8, "MODULE SUMMARY", C_HEAD, sz=11)
ws0.row_dimensions[mr].height = 22
modules_summary = [
    ("AUTH", "Authentication & Onboarding",    11, 7, 3, 1, 0, 42),
    ("DASH", "Dashboard",                        5, 2, 3, 0, 0, 18),
    ("GEN",  "Script Generator",                 8, 4, 3, 1, 0, 32),
    ("REC",  "Teleprompter & Recorder",          9, 5, 3, 1, 0, 35),
    ("HIST", "Script History",                   5, 2, 1, 2, 0, 16),
    ("CROSS","Crosspost",                        5, 3, 1, 1, 0, 16),
    ("PROF", "Profile & Settings",               7, 2, 4, 1, 0, 22),
    ("PAY",  "Payments & Subscriptions",        11, 6, 3, 2, 0, 44),
    ("NOTIF","Notifications & Email",            6, 2, 1, 2, 1, 18),
    ("MOB",  "Mobile App – iOS & Android",      12, 7, 3, 1, 1, 52),
    ("INFRA","Backend & Infrastructure",         9, 5, 3, 1, 0, 30),
    ("TEST", "Testing & QA",                     8, 4, 3, 1, 0, 32),
    ("SEC",  "Security",                         7, 6, 1, 0, 0, 24),
    ("LEGAL","Legal & Compliance",               6, 4, 2, 0, 0, 16),
    ("PERF", "Performance & Analytics",          7, 3, 3, 1, 0, 22),
]
sum_hdrs = ["Code","Module","Stories","P1","P2","P3","P4","Est. Hrs",""]
for ci, h in enumerate(sum_hdrs, start=1):
    c = ws0.cell(row=mr+1, column=ci, value=h)
    c.font = hf(10)
    c.fill = fill(C_NAV)
    c.alignment = ca()
    c.border = tb()
ws0.row_dimensions[mr+1].height = 22
total_stories = sum(m[2] for m in modules_summary)
total_hrs     = sum(m[6] for m in modules_summary)
for ri, (code, mod, tot, p1, p2, p3, p4, hrs) in enumerate(modules_summary, start=mr+2):
    row_vals = [code, mod, tot, p1, p2, p3, p4, hrs, ""]
    for ci, v in enumerate(row_vals, start=1):
        c = ws0.cell(row=ri, column=ci, value=v)
        c.border = tb()
        c.alignment = ca() if ci != 2 else ca("left")
        c.font = bf(10, bold=(ci in [1,2]))
        if ri % 2 == 0:
            c.fill = fill(C_ALT)
    ws0.row_dimensions[ri].height = 20
# totals
tr = mr + 2 + len(modules_summary)
for ci, v in enumerate(["", "TOTAL", total_stories, "", "", "", "", total_hrs, ""], start=1):
    c = ws0.cell(row=tr, column=ci, value=v)
    c.font = hf(10)
    c.fill = fill(C_NAV)
    c.alignment = ca()
    c.border = tb()
ws0.row_dimensions[tr].height = 22

cw0 = [8, 32, 10, 6, 6, 6, 6, 10]
for i, w in enumerate(cw0, 1):
    ws0.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════
# SHEET 2 — USER STORIES
# ════════════════════════════════════════════════════
ws = wb.create_sheet("User Stories")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"

col_hdrs = ["ID", "Module", "User Story", "Acceptance Criteria", "Priority",
            "Platform", "Est. Hrs", "Status", "Assigned To", "Notes"]
col_widths = [10, 22, 42, 68, 18, 14, 9, 14, 14, 22]

ws.merge_cells(f"A1:{get_column_letter(len(col_hdrs))}1")
ws["A1"] = "NUOVE — User Stories & Acceptance Criteria"
ws["A1"].font = Font(name="Arial", bold=True, size=16, color=C_WHITE)
ws["A1"].fill = fill(C_NAV)
ws["A1"].alignment = ca()
ws.row_dimensions[1].height = 36

for ci, h in enumerate(col_hdrs, 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.font = hf(10)
    c.fill = fill(C_HEAD)
    c.alignment = ca()
    c.border = tb()
ws.row_dimensions[2].height = 22

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── user stories data ──
# format: (ID, Module, User Story, Acceptance Criteria, Priority, Platform, Est_Hrs)
stories = [

  # ─── AUTH – Authentication & Onboarding ───────────────────────
  ("AUTH-01", "Auth & Onboarding",
   "As a new user, I want to sign up with my email and password so I can create an account.",
   "• Sign up form has: Name, Email, Password, Confirm Password fields\n• Password min 8 chars, must include 1 number\n• Shows inline error if email already registered\n• Shows inline error if passwords don't match\n• On success → redirect to onboarding step 1\n• Sends welcome email on signup\n• Password stored as bcrypt hash (never plain text)",
   "P1 – Must Have", "All", 4),

  ("AUTH-02", "Auth & Onboarding",
   "As a user, I want to sign in with Google so I don't have to remember a password.",
   "• 'Continue with Google' button on login and signup screen\n• Clicking opens Google OAuth popup\n• On success → new users go to onboarding, existing users go to Dashboard\n• Profile photo from Google is stored as avatar\n• If Google email matches existing account → merged automatically",
   "P1 – Must Have", "All", 5),

  ("AUTH-03", "Auth & Onboarding",
   "As an iOS user, I want to sign in with Apple ID so I can use the app per App Store requirements.",
   "• 'Sign in with Apple' button visible on iOS only\n• Apple Sign In is MANDATORY for App Store (Apple rule)\n• Returns name + email (or private relay email)\n• New Apple users go to onboarding\n• Handles 'hide my email' relay correctly",
   "P1 – Must Have", "iOS", 6),

  ("AUTH-04", "Auth & Onboarding",
   "As a returning user, I want to log in and stay logged in so I don't log in every time.",
   "• JWT access token (15 min expiry) + refresh token (30 day expiry)\n• Refresh token stored in httpOnly cookie\n• App auto-refreshes token silently in background\n• User stays logged in for 30 days without re-entering password\n• 'Remember me' checkbox on login screen",
   "P1 – Must Have", "All", 4),

  ("AUTH-05", "Auth & Onboarding",
   "As a user who forgot my password, I want to reset it via email.",
   "• 'Forgot password?' link on login screen\n• Enter email → receive reset link within 60 seconds\n• Reset link expires after 1 hour\n• New password form: min 8 chars, confirm field\n• On success → auto login → redirect to Dashboard\n• Old sessions invalidated after reset",
   "P1 – Must Have", "All", 3),

  ("AUTH-06", "Auth & Onboarding",
   "As a new user, I want an onboarding flow so I can set up my profile before using the app.",
   "• Step 1: Choose your niche (select 1–3 from: Fitness, Food, Fashion, Finance, Tech, Travel, Comedy, Education, Other)\n• Step 2: Choose primary platform (Instagram Reels / YouTube Shorts / Both)\n• Step 3: Choose language (English, Hindi, Hinglish, Tamil, Telugu, Other)\n• Step 4: 'Generate your first script' CTA\n• Progress bar showing step X of 4\n• Can skip onboarding (goes to Dashboard)\n• Selections saved to user profile in DB",
   "P1 – Must Have", "All", 5),

  ("AUTH-07", "Auth & Onboarding",
   "As a user, I want to see a splash screen when the app opens so it feels like a real native app.",
   "• App icon + 'Nuove' wordmark centred on brand gradient background\n• Visible for 1.5 seconds max\n• Does not show on subsequent app opens if already logged in (goes straight to Dashboard)\n• Required for App Store / Play Store",
   "P1 – Must Have", "iOS, Android", 2),

  ("AUTH-08", "Auth & Onboarding",
   "As a user, I want to log out securely from any device.",
   "• Logout button in Profile / Settings\n• Clears JWT tokens from storage\n• Invalidates refresh token on server\n• Redirects to login screen\n• 'Log out of all devices' option in settings",
   "P1 – Must Have", "All", 2),

  ("AUTH-09", "Auth & Onboarding",
   "As a user, I want to delete my account and all my data.",
   "• 'Delete Account' option in Settings → Account\n• Confirmation modal: type 'DELETE' to confirm\n• Deletes: user record, all scripts, recordings metadata, payment history\n• Cancels active subscription via Razorpay\n• Sends confirmation email\n• Required by DPDP Act 2023 (India)",
   "P2 – Should Have", "All", 4),

  ("AUTH-10", "Auth & Onboarding",
   "As a user on the app for the first time, I want a short product tour so I understand the features.",
   "• 3-slide carousel after onboarding (skippable)\n• Slide 1: Generate Scripts\n• Slide 2: Teleprompter & Recorder\n• Slide 3: Crosspost everywhere\n• 'Get Started' button on last slide → Dashboard\n• Never shown again after first view",
   "P2 – Should Have", "All", 3),

  ("AUTH-11", "Auth & Onboarding",
   "As a user, I want to see a proper loading/error state if login fails.",
   "• Button shows spinner while API call in progress\n• Button is disabled while loading (prevents double submit)\n• Wrong password → 'Incorrect email or password'\n• Account locked → 'Too many attempts. Try in 15 minutes.'\n• Network error → 'No internet connection. Please try again.'\n• No white screen of death under any circumstance",
   "P1 – Must Have", "All", 2),

  # ─── DASH – Dashboard ─────────────────────────────────────────
  ("DASH-01", "Dashboard",
   "As a user, I want to see a personalised greeting on my Dashboard so the app feels mine.",
   "• Shows 'Good morning/afternoon/evening, [First Name]'\n• Greeting uses current device time\n• Gradient text using brand colours\n• Sub-line shows today's date and a motivational tip (rotates daily)",
   "P2 – Should Have", "All", 2),

  ("DASH-02", "Dashboard",
   "As a user, I want to see my weekly stats on the Dashboard so I can track my progress.",
   "• Shows: Scripts Generated (this week), Recordings Made (this week), Current Streak (days)\n• Streak = consecutive days where user generated at least 1 script\n• Numbers animate on load (count up)\n• Empty state: '0 scripts this week — let's fix that!'",
   "P2 – Should Have", "All", 4),

  ("DASH-03", "Dashboard",
   "As a user, I want quick-access shortcuts on the Dashboard to my most-used features.",
   "• 4 shortcut cards: Generate Script, Record, Crosspost, Trending\n• Tapping any card navigates to that feature\n• Cards show icon + label\n• Visually matches brand design",
   "P2 – Should Have", "All", 3),

  ("DASH-04", "Dashboard",
   "As a user, I want to see my last 3 saved scripts on the Dashboard.",
   "• Shows 3 most recent scripts (truncated to 2 lines)\n• 'View All' link → Script History page\n• Empty state: 'No scripts yet. Generate your first one!'\n• Tapping a script → opens it in Script History",
   "P3 – Nice to Have", "All", 3),

  ("DASH-05", "Dashboard",
   "As a user, I want a 'content tip of the day' card on the Dashboard.",
   "• Single tip card (e.g. 'Hooks under 3 seconds get 2x more saves')\n• Rotates daily from a pool of 30 tips\n• Tip is relevant to the user's chosen niche\n• Tap to dismiss",
   "P3 – Nice to Have", "All", 3),

  # ─── GEN – Script Generator ───────────────────────────────────
  ("GEN-01", "Script Generator",
   "As a creator, I want to generate a full video script by describing my topic.",
   "• Input field: 'What's your video about?' (max 200 chars)\n• Auto-detect niche from onboarding (shown as tag, can change)\n• Auto-detect language (can change via dropdown)\n• Platform selector: Reels / Shorts / Both\n• Tone selector: Funny, Educational, Motivational, Controversial, Storytelling\n• 'Generate Script' button\n• Shows loading spinner during generation\n• Streams response (text appears word by word)",
   "P1 – Must Have", "All", 6),

  ("GEN-02", "Script Generator",
   "As a creator, I want the generated script to be structured with a Hook, Body and CTA.",
   "• Script output has 3 clearly labelled sections:\n   – Hook (first 3 seconds — attention grabber)\n   – Body (main content, 30–45 seconds)\n   – CTA (call to action, last 5 seconds)\n• Each section has a copy button\n• Character count per section\n• Total estimated video duration shown",
   "P1 – Must Have", "All", 4),

  ("GEN-03", "Script Generator",
   "As a creator, I want to save a generated script so I can use it later.",
   "• 'Save Script' button below generated output\n• Saves to DB linked to user account (not just localStorage)\n• Shows toast: 'Script saved!'\n• Saved scripts appear in Script History\n• User can name the script (optional, defaults to topic)",
   "P1 – Must Have", "All", 4),

  ("GEN-04", "Script Generator",
   "As a creator, I want to regenerate a script if I don't like the first result.",
   "• 'Regenerate' button next to Save\n• Generates a new version with same inputs\n• Previous version is NOT lost (shown as 'Version 1' vs 'Version 2')\n• Max 3 free regenerations per script (then prompts upgrade)",
   "P2 – Should Have", "All", 3),

  ("GEN-05", "Script Generator",
   "As a creator, I want to copy the full script to clipboard with one tap.",
   "• 'Copy All' button copies entire script (Hook + Body + CTA)\n• Toast confirmation: 'Copied to clipboard'\n• Works on mobile browsers (uses navigator.clipboard API with fallback)",
   "P1 – Must Have", "All", 1),

  ("GEN-06", "Script Generator",
   "As a creator, I want to go directly to the Teleprompter with my saved script.",
   "• 'Record This' button on generated script\n• Tapping saves script to sessionStorage and navigates to /record\n• Teleprompter loads with the script pre-filled\n• No need to copy-paste manually",
   "P1 – Must Have", "All", 2),

  ("GEN-07", "Script Generator",
   "As a free-tier user, I want to know how many script generations I have left.",
   "• Free plan: 5 generations per month\n• Counter shown: 'X of 5 generations used this month'\n• On hitting limit: modal appears with upgrade prompt\n• Counter resets on 1st of each month\n• Pro plan: unlimited generations",
   "P2 – Should Have", "All", 4),

  ("GEN-08", "Script Generator",
   "As a creator, I want to generate scripts in multiple Indian languages.",
   "• Language options: English, Hindi, Hinglish, Tamil, Telugu, Kannada, Bengali\n• Language selector visible in generate form\n• Script generated in selected language\n• Language saved to user preference automatically",
   "P2 – Should Have", "All", 3),

  # ─── REC – Teleprompter & Recorder ───────────────────────────
  ("REC-01", "Teleprompter & Recorder",
   "As a creator, I want to see my script scroll automatically while recording so I don't need a second device.",
   "• Script text displayed large, centred on screen\n• Auto-scroll starts when recording begins\n• Auto-scroll pauses when recording is paused\n• Speed control: 7 levels (Very Slow to Fast)\n• Font size control: Small / Medium / Large\n• Text is white on dark background for readability",
   "P1 – Must Have", "All", 5),

  ("REC-02", "Teleprompter & Recorder",
   "As a creator, I want to record video using my front camera while reading the teleprompter.",
   "• Requests camera + microphone permission on first use\n• Front camera selected by default\n• Camera feed visible behind the script text\n• Permission denied → shows 'Please allow camera access in Settings'\n• Works on iOS Safari, Android Chrome, and web",
   "P1 – Must Have", "All", 6),

  ("REC-03", "Teleprompter & Recorder",
   "As a creator, I want a countdown before recording starts so I can get ready.",
   "• 3-2-1 countdown displayed large before recording starts\n• Countdown sound (optional, can mute)\n• Recording indicator (red dot + 'REC' label) visible during recording",
   "P1 – Must Have", "All", 2),

  ("REC-04", "Teleprompter & Recorder",
   "As a creator, I want to pause and resume recording.",
   "• Pause button visible during recording\n• Pausing freezes teleprompter scroll\n• Resuming continues from same position\n• Recording timer pauses when paused\n• Final video is continuous (no gap at pause point)",
   "P2 – Should Have", "All", 4),

  ("REC-05", "Teleprompter & Recorder",
   "As a creator, I want to preview and download my recording when done.",
   "• After stopping, show video playback\n• 'Download' button saves video to device\n• Download filename: nuove-recording.mp4\n• iOS: shows share sheet (save to Files / Photos)\n• Android: saves to Downloads folder",
   "P1 – Must Have", "All", 4),

  ("REC-06", "Teleprompter & Recorder",
   "As a creator, I want to manually type or paste a script into the teleprompter.",
   "• Text area to enter script manually\n• Pre-filled if arriving from Generate page\n• Can edit text directly in the teleprompter setup screen\n• 'Clear' button to remove all text",
   "P1 – Must Have", "All", 2),

  ("REC-07", "Teleprompter & Recorder",
   "As a creator on iOS, I want my recording to work in the correct format.",
   "• iOS Safari does not support webm — use video/mp4 or video/quicktime\n• Detect browser/OS and select correct codec automatically\n• Test on: iPhone 12, 13, 14 (iOS 15, 16, 17)\n• File plays correctly in iPhone Camera Roll",
   "P1 – Must Have", "iOS", 5),

  ("REC-08", "Teleprompter & Recorder",
   "As a creator, I want the recorder to work in fullscreen on mobile.",
   "• Entering record mode requests fullscreen on Android (Fullscreen API)\n• On iOS: full viewport height (no browser chrome)\n• Teleprompter text fills screen\n• Back button / swipe exits recording safely (with confirmation if recording active)",
   "P2 – Should Have", "iOS, Android", 3),

  ("REC-09", "Teleprompter & Recorder",
   "As a creator, I want to switch between front and back camera.",
   "• Camera switch button (flip icon) visible on recorder\n• Switches camera without stopping recording session\n• Camera restarts with new facing mode",
   "P3 – Nice to Have", "iOS, Android", 3),

  # ─── HIST – Script History ────────────────────────────────────
  ("HIST-01", "Script History",
   "As a creator, I want to see all my saved scripts in one place.",
   "• List of all saved scripts, newest first\n• Each item shows: Title, Date saved, First 2 lines of script\n• 'Load More' / pagination for >20 scripts\n• Empty state: 'No scripts yet. Generate your first one!' + CTA button",
   "P1 – Must Have", "All", 3),

  ("HIST-02", "Script History",
   "As a creator, I want to open and read a saved script in full.",
   "• Tap script → opens full view (modal or new screen)\n• Full Hook / Body / CTA sections visible\n• Copy button per section and 'Copy All'\n• 'Record This' button → opens Teleprompter with script",
   "P1 – Must Have", "All", 3),

  ("HIST-03", "Script History",
   "As a creator, I want to delete scripts I no longer need.",
   "• Delete button (trash icon) on each script\n• Confirmation: 'Delete this script? This cannot be undone.'\n• On confirm → removed from list with animation\n• Synced to DB (deletion is permanent)",
   "P2 – Should Have", "All", 2),

  ("HIST-04", "Script History",
   "As a creator, I want to search my saved scripts.",
   "• Search bar at top of Scripts page\n• Filters list in real time by script title or content\n• No results state: 'No scripts match your search'",
   "P3 – Nice to Have", "All", 3),

  ("HIST-05", "Script History",
   "As a creator, I want to share a script with someone via a link.",
   "• 'Share' button on each script\n• Generates a public read-only URL: nuove.in/s/[unique-id]\n• Anyone with link can view (no login required)\n• Shared view is read-only — no edit/delete\n• 'Copy Link' copies URL to clipboard",
   "P3 – Nice to Have", "All", 5),

  # ─── CROSS – Crosspost ───────────────────────────────────────
  ("CROSS-01", "Crosspost",
   "As a creator, I want to write a caption once and format it for all platforms.",
   "• Input: paste your script or video topic\n• Select platforms: Instagram, YouTube, TikTok, LinkedIn, Twitter/X\n• AI generates platform-optimised captions for each selected platform\n• Each caption follows platform best practices (length, hashtag count, tone)",
   "P1 – Must Have", "All", 5),

  ("CROSS-02", "Crosspost",
   "As a creator, I want the captions to include relevant hashtags.",
   "• Each caption includes 5–15 relevant hashtags\n• Hashtags are researched per platform norms\n• Instagram: 10–15 hashtags\n• YouTube: 3–5 hashtags in description\n• TikTok: 5–8 hashtags\n• Hashtags can be edited by user",
   "P1 – Must Have", "All", 3),

  ("CROSS-03", "Crosspost",
   "As a creator, I want to copy any platform caption with one tap.",
   "• 'Copy' button per platform caption\n• Toast: 'Copied!'\n• Works on mobile (navigator.clipboard with textarea fallback)\n• Entire caption + hashtags copied together",
   "P1 – Must Have", "All", 1),

  ("CROSS-04", "Crosspost",
   "As a creator, I want to generate captions in different languages.",
   "• Language selector in Crosspost page\n• Options: English, Hindi, Hinglish, Tamil, Telugu, Kannada\n• All platform captions generated in selected language\n• Language preference remembered from last use",
   "P2 – Should Have", "All", 3),

  ("CROSS-05", "Crosspost",
   "As a creator, I want to regenerate a specific platform's caption if I don't like it.",
   "• 'Regenerate' icon button on each platform's caption card\n• Only regenerates that platform's caption (others unchanged)\n• Shows spinner on that card only\n• Previous version replaced on success",
   "P3 – Nice to Have", "All", 3),

  # ─── PROF – Profile & Settings ───────────────────────────────
  ("PROF-01", "Profile & Settings",
   "As a user, I want to view and edit my profile information.",
   "• Shows: Name, Email, Profile Photo, Niche, Platform preference, Language\n• Edit Name: inline text field\n• Edit profile photo: tap photo → image picker\n• Email change requires current password confirmation\n• Save button → updates DB → shows success toast",
   "P1 – Must Have", "All", 4),

  ("PROF-02", "Profile & Settings",
   "As a user, I want to change my password from within the app.",
   "• Settings → Security → Change Password\n• Fields: Current Password, New Password, Confirm New Password\n• Validates current password before allowing change\n• All existing sessions logged out after change\n• Success toast + email notification sent",
   "P2 – Should Have", "All", 3),

  ("PROF-03", "Profile & Settings",
   "As a user, I want to see my subscription plan and usage in Profile.",
   "• Shows current plan: Free / Pro / Creator\n• Shows usage: X of 5 script generations used\n• 'Upgrade' button → Pricing page\n• 'Manage Subscription' → Billing page (for paid users)",
   "P1 – Must Have", "All", 3),

  ("PROF-04", "Profile & Settings",
   "As a user, I want to control notification preferences.",
   "• Toggle: Email notifications (weekly tips, feature updates)\n• Toggle: Push notifications (if PWA/native app)\n• Toggle: Marketing emails\n• Saved to user preferences in DB",
   "P3 – Nice to Have", "All", 3),

  ("PROF-05", "Profile & Settings",
   "As a user, I want to toggle between dark and light mode.",
   "• Toggle in Settings or Profile page\n• Defaults to system preference (prefers-color-scheme)\n• Preference saved to localStorage\n• All pages must look correct in both modes",
   "P2 – Should Have", "All", 3),

  ("PROF-06", "Profile & Settings",
   "As a user, I want to see the app version, Terms, Privacy Policy and support links.",
   "• Settings → About section\n• App version number\n• Links: Terms & Conditions, Privacy Policy, Contact Support\n• Links open in in-app browser (not external browser)\n• 'Rate this App' link (iOS: App Store, Android: Play Store)",
   "P2 – Should Have", "All", 2),

  ("PROF-07", "Profile & Settings",
   "As a user, I want to change my content niche and language preference.",
   "• Settings → Preferences\n• Niche multi-select (same options as onboarding)\n• Language dropdown\n• Platform preference: Reels / Shorts / Both\n• Changes apply to all future script generations",
   "P2 – Should Have", "All", 2),

  # ─── PAY – Payments & Subscriptions ─────────────────────────
  ("PAY-01", "Payments & Subscriptions",
   "As a user, I want to see a pricing page with clear plan comparison.",
   "• 3 tiers: Free, Pro, Creator\n• Feature comparison table (what's included in each)\n• Monthly / Annual toggle (Annual = 2 months free)\n• 'Most Popular' badge on Pro plan\n• CTA: 'Start Free', 'Go Pro', 'Go Creator'\n• INR pricing shown (₹)",
   "P1 – Must Have", "All", 5),

  ("PAY-02", "Payments & Subscriptions",
   "As a user, I want to subscribe using Razorpay with my UPI, card or net banking.",
   "• 'Subscribe' button opens Razorpay checkout\n• Razorpay checkout handles: UPI, Cards, Net Banking, Wallets\n• Live Razorpay keys configured (not test keys)\n• On payment success → plan activated immediately\n• On payment failure → friendly error message, not a white screen",
   "P1 – Must Have", "All", 5),

  ("PAY-03", "Payments & Subscriptions",
   "As a subscribed user, my features unlock immediately after payment.",
   "• Razorpay webhook: payment.captured → update user plan in DB\n• Frontend polls or receives socket event to update plan\n• Feature gates re-check user plan in real time\n• No manual refresh needed to see unlocked features",
   "P1 – Must Have", "All", 5),

  ("PAY-04", "Payments & Subscriptions",
   "As a subscribed user, I want to see my billing history and invoices.",
   "• Settings → Billing → History\n• List of payments: Date, Amount, Plan, Status\n• 'Download Invoice' (PDF) per transaction\n• Invoice includes GSTIN (required for Indian businesses)",
   "P2 – Should Have", "All", 4),

  ("PAY-05", "Payments & Subscriptions",
   "As a subscribed user, I want to cancel my subscription from within the app.",
   "• Settings → Billing → Cancel Subscription\n• Confirmation modal explaining what they'll lose\n• On confirm → calls Razorpay cancel API\n• Plan remains active until end of billing period\n• Confirmation email sent\n• Can re-subscribe at any time",
   "P2 – Should Have", "All", 3),

  ("PAY-06", "Payments & Subscriptions",
   "As a free user, I want a clear prompt to upgrade when I hit my usage limit.",
   "• Modal appears when free limit is reached (e.g. 5 scripts/month)\n• Explains what's included in Pro\n• 'Upgrade Now' CTA\n• 'Maybe Later' dismisses\n• Counter visible on Generate page: 'X of 5 used'",
   "P1 – Must Have", "All", 3),

  ("PAY-07", "Payments & Subscriptions",
   "As a user, I want to receive a payment receipt by email.",
   "• Email sent automatically on successful payment\n• Contains: Plan name, Amount, Invoice number, Date, Support link\n• Sent from noreply@nuove.in\n• GST invoice attached as PDF",
   "P1 – Must Have", "All", 3),

  ("PAY-08", "Payments & Subscriptions",
   "As a subscribed user, I want to receive a reminder before my subscription renews.",
   "• Email sent 3 days before renewal date\n• Contains: Amount, Date, 'Manage Subscription' link\n• Push notification also sent (if enabled)\n• If payment method will expire → extra warning email",
   "P2 – Should Have", "All", 2),

  ("PAY-09", "Payments & Subscriptions",
   "As a user, I want to be able to use a promo/referral code for a discount.",
   "• Promo code field in checkout\n• Backend validates code → applies discount\n• Shows: 'Code applied! 30% off'\n• Codes have expiry date and usage limit\n• Invalid code shows 'Invalid or expired code'",
   "P3 – Nice to Have", "All", 4),

  ("PAY-10", "Payments & Subscriptions",
   "As a subscribed user on iOS, I want to subscribe via Apple In-App Purchase.",
   "• Apple requires IAP for subscriptions sold via iOS app (their rule)\n• Implement StoreKit 2 via Capacitor plugin\n• Plan pricing must match equivalent Razorpay pricing (Apple takes 15–30%)\n• Apple subscription and Razorpay subscription must sync to same user plan\n• NOTE: This is complex — do this after initial launch",
   "P3 – Nice to Have", "iOS", 8),

  ("PAY-11", "Payments & Subscriptions",
   "As a user on Android, I want to subscribe via Google Play Billing.",
   "• Google requires Play Billing for in-app purchases sold via Play Store\n• Implement via Capacitor Purchases plugin\n• Sync Google subscription status to backend\n• NOTE: Like Apple IAP — do after initial launch",
   "P3 – Nice to Have", "Android", 8),

  # ─── NOTIF – Notifications & Email ───────────────────────────
  ("NOTIF-01", "Notifications & Email",
   "As a new user, I want to receive a welcome email after signing up.",
   "• Email sent within 60 seconds of signup\n• Subject: 'Welcome to Nuove 🎬'\n• Contains: User's name, Quick start guide, Link to Dashboard\n• HTML email with brand colours\n• Does NOT land in spam (SPF/DKIM configured for nuove.in)",
   "P1 – Must Have", "All", 3),

  ("NOTIF-02", "Notifications & Email",
   "As a user, I want to receive a weekly content tip email.",
   "• Sent every Monday at 9 AM IST\n• One actionable tip related to user's niche\n• 'Generate a script now' CTA button in email\n• Unsubscribe link at bottom (required by law)\n• Sent via Resend.com or SendGrid",
   "P3 – Nice to Have", "All", 4),

  ("NOTIF-03", "Notifications & Email",
   "As a user, I want to receive an email when a new feature is released.",
   "• Feature announcement email template\n• Sent to all active users\n• Includes GIF/screenshot of new feature\n• 'Try it now' CTA\n• Unsubscribe option",
   "P3 – Nice to Have", "All", 2),

  ("NOTIF-04", "Notifications & Email",
   "As a mobile user, I want to receive push notifications for important updates.",
   "• Request push permission after user generates first script\n• Notifications: 'Your weekly tip is ready!', payment reminders, streak alerts\n• Push via Firebase Cloud Messaging (FCM)\n• Works on Android (native) and iOS (via Capacitor)\n• User can disable in Settings",
   "P2 – Should Have", "iOS, Android", 5),

  ("NOTIF-05", "Notifications & Email",
   "As a user, I want in-app notifications for key events.",
   "• Bell icon in top nav shows unread count badge\n• Notification types: Script saved, Payment received, Streak milestone, New feature\n• Tap notification → navigates to relevant screen\n• Mark all as read button\n• Notifications stored in DB (persist across sessions)",
   "P3 – Nice to Have", "All", 4),

  ("NOTIF-06", "Notifications & Email",
   "As a user, I want a streak notification if I haven't opened the app in 2 days.",
   "• Push notification after 48h of no activity: 'Your streak is at risk!'\n• Shown only if user had a streak > 2 days\n• Sent at 6 PM IST (prime mobile usage time in India)\n• Can disable in notification settings",
   "P4 – Future", "iOS, Android", 3),

  # ─── MOB – Mobile App ─────────────────────────────────────────
  ("MOB-01", "Mobile App",
   "As a user, I want the app wrapped as a native iOS app for App Store submission.",
   "• Use Capacitor to wrap the React/Vite app\n• App opens in fullscreen (no browser chrome)\n• iOS minimum target: iOS 15\n• App icons: all required sizes (20x20 to 1024x1024)\n• App name: 'Nuove'\n• Bundle ID: in.nuove.app\n• Passes Apple App Review Guidelines",
   "P1 – Must Have", "iOS", 8),

  ("MOB-02", "Mobile App",
   "As a user, I want the app wrapped as a native Android app for Play Store.",
   "• Use Capacitor for Android build\n• Android minimum target: API 26 (Android 8)\n• App icons: all required densities\n• Package name: in.nuove.app\n• Signed APK / AAB for Play Store\n• Passes Google Play policies",
   "P1 – Must Have", "Android", 8),

  ("MOB-03", "Mobile App",
   "As a mobile user, I want the app to look correct on all screen sizes.",
   "• Test on: iPhone SE (375px), iPhone 14 (390px), iPhone 14 Pro Max (430px)\n• Test on: Samsung Galaxy A series (360px wide), Pixel (412px)\n• No horizontal scrolling on any page\n• No text overflow or clipping\n• Bottom nav is thumb-friendly (min 56px height, min 44px tap target)",
   "P1 – Must Have", "iOS, Android", 6),

  ("MOB-04", "Mobile App",
   "As a mobile user, I want the app to handle soft keyboard without breaking the layout.",
   "• When soft keyboard opens, layout adjusts correctly\n• Input fields scroll above keyboard (not hidden behind it)\n• Use window.visualViewport API to detect keyboard height\n• Test on iOS Safari and Android Chrome\n• Modal forms do not get stuck behind keyboard",
   "P1 – Must Have", "iOS, Android", 4),

  ("MOB-05", "Mobile App",
   "As a mobile user, I want proper permission dialogs for camera and microphone.",
   "• First launch → native permission dialog (Capacitor handles this)\n• If denied → show in-app guide: 'Go to Settings → Nuove → Allow Camera'\n• Include a screenshot/illustration of settings location\n• Permission check on every entry to Record page\n• Graceful degradation if mic denied (record video only with no audio)",
   "P1 – Must Have", "iOS, Android", 4),

  ("MOB-06", "Mobile App",
   "As an iOS user, I want to save my recording directly to Photos.",
   "• After recording, offer 'Save to Photos' in addition to Download\n• Uses Capacitor Filesystem + Photos plugin\n• Requests photo library write permission\n• Success toast: 'Saved to Photos!'",
   "P2 – Should Have", "iOS", 3),

  ("MOB-07", "Mobile App",
   "As a mobile user, I want the app to work when I switch between apps (background/foreground).",
   "• App state preserved when backgrounded\n• If recording is active and user backgrounds app → pause recording + show alert on return\n• Session token not lost on backgrounding\n• Teleprompter position remembered on return",
   "P2 – Should Have", "iOS, Android", 4),

  ("MOB-08", "Mobile App",
   "As a user, I want the app to load fast on mobile (including slow 4G).",
   "• Time to Interactive < 3 seconds on 4G\n• React.lazy() used for all page components\n• Total JS bundle < 300KB gzipped\n• Images in WebP format\n• Fonts preloaded in <head>",
   "P1 – Must Have", "iOS, Android", 5),

  ("MOB-09", "Mobile App",
   "As a user, I want smooth animations on mobile (no jank).",
   "• All animations run at 60fps on mid-range Android (2GB RAM)\n• Avoid animating layout properties (width/height) — use transform/opacity only\n• Reduce motion respected (prefers-reduced-motion media query)\n• No animation flicker on page transitions",
   "P2 – Should Have", "iOS, Android", 3),

  ("MOB-10", "Mobile App",
   "As a user, I want the app to be submittable to App Store and Play Store.",
   "APP STORE requirements:\n• Privacy Policy URL in App Store Connect\n• App Store screenshots: 6.5\" and 5.5\" iPhone sizes\n• App description (170 char subtitle + full description)\n• Age rating: 4+\n• No use of private APIs\n• Apple Sign In implemented (AUTH-03)\n\nPLAY STORE requirements:\n• Feature graphic (1024x500px)\n• 2 screenshots minimum per device type\n• Data safety section completed\n• Target API 33+ (Android 13)",
   "P1 – Must Have", "iOS, Android", 6),

  ("MOB-11", "Mobile App",
   "As a user, I want the app icon and splash screen to match the Nuove brand.",
   "• App icon: Nuove clapperboard logo on gradient background\n• No text in icon (Apple guideline)\n• Splash screen: brand gradient + Nuove wordmark centred\n• Splash auto-hides after 1.5 seconds\n• Dark/light mode app icon variants",
   "P1 – Must Have", "iOS, Android", 3),

  ("MOB-12", "Mobile App",
   "As an Android user, I want to install the app via a direct APK link if Play Store review is pending.",
   "• Build signed APK\n• Host on nuove.in/download/nuove.apk\n• Landing page with QR code for easy download\n• Note: This is for beta testing only. Production uses Play Store.",
   "P2 – Should Have", "Android", 2),

  # ─── INFRA – Backend & Infrastructure ────────────────────────
  ("INFRA-01", "Infrastructure",
   "As the product owner, I want the backend hosted on a paid Railway plan so it never sleeps.",
   "• Upgrade Railway from trial → Hobby ($5/mo) or Pro\n• Backend responds in < 500ms (no cold start delay)\n• Health check endpoint: GET /health returns 200\n• Railway auto-restarts on crash\n• FRONTEND_URL env var set to https://nuove.in",
   "P1 – Must Have", "All", 1),

  ("INFRA-02", "Infrastructure",
   "As the product owner, I want a PostgreSQL database storing all user data.",
   "• Add Railway PostgreSQL plugin OR use Supabase free tier\n• Tables: users, scripts, recordings_meta, subscriptions, payments, notifications\n• DATABASE_URL set in Railway env vars\n• DB backups enabled (daily minimum)\n• Connection pooling configured (PgBouncer or pg pool)",
   "P1 – Must Have", "All", 5),

  ("INFRA-03", "Infrastructure",
   "As the product owner, I want all app secrets in environment variables, not in code.",
   "• Audit: no API keys in any committed file\n• Railway env vars: DATABASE_URL, JWT_SECRET, ANTHROPIC_API_KEY, RAZORPAY keys, RESEND key\n• Vercel env vars: VITE_API_URL\n• .env.example updated with correct placeholders\n• Run npm audit — 0 critical vulnerabilities",
   "P1 – Must Have", "All", 2),

  ("INFRA-04", "Infrastructure",
   "As the product owner, I want CI/CD so the app deploys automatically on code push.",
   "• GitHub Actions: run npm test + npm run lint on every PR\n• Merge to main → auto-deploy to Railway (backend) and Vercel (frontend)\n• PR blocked from merge if tests fail\n• Deployment takes < 3 minutes",
   "P2 – Should Have", "All", 5),

  ("INFRA-05", "Infrastructure",
   "As the product owner, I want uptime monitoring so I know if the app goes down.",
   "• UptimeRobot (free): ping /health every 5 minutes\n• Alert via email + Slack if down for > 2 minutes\n• Also prevents Railway free tier sleep (if still on free)\n• Public status page (optional): status.nuove.in",
   "P1 – Must Have", "All", 1),

  ("INFRA-06", "Infrastructure",
   "As the product owner, I want error tracking to catch bugs in production.",
   "• Sentry.io added to both frontend and backend (free tier)\n• Source maps uploaded for readable stack traces\n• Alerts for new errors sent to Slack/email\n• Errors tagged with userId for faster debugging",
   "P2 – Should Have", "All", 4),

  ("INFRA-07", "Infrastructure",
   "As the product owner, I want file storage for user uploads and recordings metadata.",
   "• Cloudflare R2 (free 10GB) or AWS S3 for video/image storage\n• Videos NOT stored long-term — only metadata stored in DB\n• Profile photos stored in R2 with public read URL\n• File size limit: 50MB per upload",
   "P2 – Should Have", "All", 4),

  ("INFRA-08", "Infrastructure",
   "As the product owner, I want transactional email working from nuove.in domain.",
   "• Resend.com free tier (3,000 emails/month)\n• DNS records: SPF, DKIM, DMARC configured for nuove.in\n• Verify domain in Resend dashboard\n• Test: send email from noreply@nuove.in → should not land in spam\n• EMAIL_FROM=Nuove <noreply@nuove.in>",
   "P1 – Must Have", "All", 2),

  ("INFRA-09", "Infrastructure",
   "As the product owner, I want the backend CORS configured correctly for nuove.in.",
   "• ALLOWED_ORIGINS in app.js includes: https://nuove.in, https://www.nuove.in\n• Also allows all *.vercel.app for preview deployments\n• All API calls from nuove.in return correct CORS headers\n• No CORS errors in browser console on production",
   "P1 – Must Have", "All", 1),

  # ─── TEST – Testing & QA ──────────────────────────────────────
  ("TEST-01", "Testing & QA",
   "As the intern, I want a manual QA checklist to test every feature before release.",
   "• Create test cases for: Login, Signup, Generate Script, Save Script, Record, Download, Crosspost, Billing, Settings\n• Each test case has: Steps to reproduce, Expected result, Pass/Fail column\n• Run checklist before every production deployment\n• Document in Notion or Google Sheets",
   "P1 – Must Have", "All", 5),

  ("TEST-02", "Testing & QA",
   "As the intern, I want the critical API routes to have automated tests.",
   "• Use Jest + Supertest\n• Test: POST /auth/signup, POST /auth/login, POST /generate, GET /scripts, POST /payments/webhook\n• Tests run in CI pipeline\n• Minimum: happy path + 1 error case per route",
   "P2 – Should Have", "All", 8),

  ("TEST-03", "Testing & QA",
   "As the product owner, I want the app tested on real devices before launch.",
   "• Test on: iPhone 12/13/14 (iOS 15, 16, 17)\n• Test on: Samsung Galaxy A32 / Redmi Note (Android 10, 12)\n• Test on: iPad (tablet layout acceptable)\n• Document any device-specific bugs found\n• All P1 bugs fixed before launch",
   "P1 – Must Have", "iOS, Android", 8),

  ("TEST-04", "Testing & QA",
   "As the intern, I want to run Lighthouse and fix performance scores before App Store submission.",
   "• Run Lighthouse on nuove.in on mobile preset\n• Target scores: Performance > 85, Accessibility > 90, SEO > 90, Best Practices > 90\n• Fix all issues flagged as 'Needs Improvement' or 'Poor'\n• Run in GitHub Actions CI to catch regressions",
   "P2 – Should Have", "All", 4),

  ("TEST-05", "Testing & QA",
   "As the intern, I want to test the full payment flow end-to-end.",
   "• Test with Razorpay test cards (4111 1111 1111 1111)\n• Flow: Click Upgrade → Razorpay opens → Pay → Plan activates\n• Test payment failure (declined card)\n• Test subscription cancellation\n• Test webhook delivery (use Razorpay webhook simulator)\n• Verify email receipt arrives",
   "P1 – Must Have", "All", 5),

  ("TEST-06", "Testing & QA",
   "As the intern, I want to test the recording feature across all major browsers.",
   "• Test on: iOS Safari 15+, iOS Chrome, Android Chrome 90+, Desktop Chrome, Desktop Firefox, Desktop Safari\n• Camera opens correctly\n• Recording starts/stops correctly\n• Downloaded file plays in native player\n• Document any browser-specific issues",
   "P1 – Must Have", "iOS, Android", 5),

  ("TEST-07", "Testing & QA",
   "As the product owner, I want the app to handle network errors gracefully.",
   "• Simulate offline mode (DevTools → Offline)\n• Each API call failure shows a user-friendly error (not console error)\n• Retry button shown where applicable\n• No white screen of death under any network condition\n• Test slow 3G preset in Chrome DevTools",
   "P2 – Should Have", "All", 3),

  ("TEST-08", "Testing & QA",
   "As the intern, I want to run accessibility audit before submission.",
   "• Run axe-core or Lighthouse Accessibility audit on all pages\n• Fix all 'Critical' and 'Serious' violations\n• All images have alt text\n• All form inputs have labels\n• Minimum contrast ratio 4.5:1 for body text\n• Keyboard navigation works on all interactive elements",
   "P2 – Should Have", "All", 5),

  # ─── SEC – Security ───────────────────────────────────────────
  ("SEC-01", "Security",
   "As the product owner, I want rate limiting on all API routes to prevent abuse.",
   "• Use express-rate-limit\n• /auth/login: max 5 requests per 15 minutes per IP\n• /auth/signup: max 3 per hour per IP\n• /generate: max 10 per minute per user\n• Returns 429 with message: 'Too many requests. Please wait.'\n• Prevents AI API cost explosion from abuse",
   "P1 – Must Have", "All", 3),

  ("SEC-02", "Security",
   "As the product owner, I want all user inputs validated and sanitised on the backend.",
   "• Use Zod or Joi for schema validation on every POST/PUT route\n• Validate: email format, password strength, string lengths\n• Strip HTML tags from text inputs (XSS prevention)\n• Return 400 with clear error message for invalid input\n• Never pass unsanitised input to the AI API",
   "P1 – Must Have", "All", 4),

  ("SEC-03", "Security",
   "As the product owner, I want security headers on all API responses.",
   "• Use helmet.js middleware\n• Headers: HSTS, X-Frame-Options, X-Content-Type-Options, Referrer-Policy\n• No server version headers\n• CSP configured to allow nuove.in and CDN sources only",
   "P1 – Must Have", "All", 2),

  ("SEC-04", "Security",
   "As the product owner, I want Razorpay webhook signature verified before processing.",
   "• Every incoming Razorpay webhook must verify HMAC-SHA256 signature\n• Use razorpay.webhooks.verify(body, signature, secret)\n• Reject unverified webhooks with 400\n• Log all webhook events\n• Never process payment without signature verification",
   "P1 – Must Have", "All", 2),

  ("SEC-05", "Security",
   "As the product owner, I want no sensitive data exposed in frontend bundle.",
   "• Audit Vite build output — no Anthropic/OpenAI keys in bundle\n• No Razorpay secret key in frontend\n• All AI calls go through backend proxy\n• Environment variables with VITE_ prefix are exposed to frontend — audit all of them",
   "P1 – Must Have", "All", 2),

  ("SEC-06", "Security",
   "As the product owner, I want brute-force protection on the login endpoint.",
   "• After 5 failed login attempts: lock account for 15 minutes\n• Show: 'Account locked. Try again in 15 minutes or reset your password.'\n• Implemented via per-email attempt counter in DB or Redis\n• Admin can manually unlock accounts",
   "P1 – Must Have", "All", 3),

  ("SEC-07", "Security",
   "As the product owner, I want all dependencies audited for vulnerabilities.",
   "• Run npm audit on both frontend and backend\n• Fix all HIGH and CRITICAL severity vulnerabilities\n• Add npm audit to CI pipeline — fail build on HIGH+\n• Review before every App Store submission",
   "P2 – Should Have", "All", 2),

  # ─── LEGAL – Legal & Compliance ──────────────────────────────
  ("LEGAL-01", "Legal & Compliance",
   "As the product owner, I need a Terms & Conditions page before accepting payments.",
   "• Route: nuove.in/terms\n• Covers: usage policy, AI-generated content ownership, refund policy, prohibited use, account termination\n• Must be live and linked from: signup page, checkout page, Settings → About\n• Required by Razorpay live account activation\n• Consult a lawyer or use a reputable SaaS T&C template",
   "P1 – Must Have", "All", 3),

  ("LEGAL-02", "Legal & Compliance",
   "As the product owner, I need a Privacy Policy compliant with DPDP Act 2023.",
   "• Route: nuove.in/privacy\n• Covers: what data is collected, how it's used, storage location, user rights, deletion process\n• DPDP Act 2023 (India) compliance: user consent, data principal rights\n• Linked from: signup page, Settings → About\n• Must state: Nuove does NOT sell user data",
   "P1 – Must Have", "All", 3),

  ("LEGAL-03", "Legal & Compliance",
   "As a user, I want to give clear consent to data collection when I sign up.",
   "• Signup page has: checkbox 'I agree to Terms & Conditions and Privacy Policy'\n• Checkbox is NOT pre-checked\n• Links open in modal or new tab\n• Cannot submit form without checking the box\n• Consent timestamp stored in DB",
   "P1 – Must Have", "All", 2),

  ("LEGAL-04", "Legal & Compliance",
   "As the product owner, I need an AI content disclosure visible to users.",
   "• Disclaimer visible on Script Generator output: 'Scripts are AI-generated. Please review before posting.'\n• Brief statement in T&C about AI-generated content\n• Does not imply Nuove owns copyright to generated scripts\n• Required by emerging AI content regulations",
   "P2 – Should Have", "All", 1),

  ("LEGAL-05", "Legal & Compliance",
   "As the product owner, I want to register the Nuove trademark and verify GST.",
   "• Check trademark: IP India search for 'Nuove' in Class 9 (software) and Class 42 (SaaS)\n• File trademark application (~₹4,500) if clear\n• GSTIN required for Razorpay live account and GST invoices\n• Register company (Pvt Ltd or LLP) if not already done\n• This is a business task — involves a CA and lawyer",
   "P1 – Must Have", "All", 0),

  ("LEGAL-06", "Legal & Compliance",
   "As the product owner, I want cookie consent for European users.",
   "• Add CookieYes or Cookiebot banner (free tier)\n• Only essential cookies set before consent\n• Analytics/marketing cookies only after explicit accept\n• 'Reject All' option must be equally prominent as 'Accept'\n• Required for GDPR compliance",
   "P2 – Should Have", "All", 2),

  # ─── PERF – Performance & Analytics ──────────────────────────
  ("PERF-01", "Performance & Analytics",
   "As the product owner, I want Google Analytics 4 tracking user behaviour.",
   "• Add GA4 to Vercel project\n• Track custom events: script_generated, recording_started, recording_completed, subscription_started, subscription_cancelled\n• Page view tracking automatic\n• Filter out localhost traffic\n• Connect GA4 to Google Search Console",
   "P1 – Must Have", "All", 3),

  ("PERF-02", "Performance & Analytics",
   "As the product owner, I want to see where users drop off in the product funnel.",
   "• Define funnel in GA4 or Posthog: Signup → Generate Script → Save Script → Record → Share/Crosspost → Subscribe\n• Weekly review of funnel drop-off rates\n• Alert if conversion rate drops > 20% week-over-week",
   "P2 – Should Have", "All", 3),

  ("PERF-03", "Performance & Analytics",
   "As the product owner, I want Hotjar or Microsoft Clarity for session recordings.",
   "• Microsoft Clarity (free, no data limit)\n• Heatmaps on: Dashboard, Generate, Record, Pricing pages\n• Session recordings enabled\n• Review weekly to find UX problems\n• Add after launch (not needed pre-launch)",
   "P3 – Nice to Have", "All", 2),

  ("PERF-04", "Performance & Analytics",
   "As the product owner, I want the frontend bundle optimised for fast mobile load.",
   "• React.lazy() + Suspense for all page-level components\n• Run: npx vite-bundle-analyzer\n• Remove unused imports\n• Replace moment.js with dayjs (10x smaller)\n• Total bundle: < 300KB gzipped\n• First Contentful Paint < 1.5s on 4G",
   "P1 – Must Have", "All", 4),

  ("PERF-05", "Performance & Analytics",
   "As the product owner, I want AI responses to stream instead of waiting for full response.",
   "• Use Anthropic streaming API\n• Frontend renders text word-by-word as it arrives\n• Shows spinner only for first 500ms, then text starts appearing\n• User perceives faster response even if total time is same\n• Implement in /generate endpoint",
   "P2 – Should Have", "All", 5),

  ("PERF-06", "Performance & Analytics",
   "As the product owner, I want a public landing page to drive organic signups.",
   "• Route: nuove.in (before login redirect)\n• Sections: Hero, Features, How it works, Pricing, Testimonials, CTA\n• SEO meta tags: title, description, og:image, og:title\n• Mobile-first design\n• Page load < 2s\n• 'Get Started Free' CTA → signup page",
   "P1 – Must Have", "All", 10),

  ("PERF-07", "Performance & Analytics",
   "As the product owner, I want sitemap.xml and robots.txt for SEO.",
   "• /sitemap.xml lists all public pages (landing, features, pricing, terms, privacy)\n• /robots.txt allows all crawlers on public pages, disallows /api and /dashboard\n• Submit sitemap to Google Search Console\n• Add canonical tags to prevent duplicate content",
   "P2 – Should Have", "All", 2),
]

STATUS_DEFAULT = "Not Started"
ASSIGNED_DEFAULT = "Intern"

row_idx = 3
prev_module = None

for (sid, module, story, criteria, priority, platform, hrs) in stories:
    # Section header row when module changes
    if module != prev_module:
        ws.merge_cells(f"A{row_idx}:{get_column_letter(len(col_hdrs))}{row_idx}")
        ws[f"A{row_idx}"] = f"  ▸  {module.upper()}"
        ws[f"A{row_idx}"].font = Font(name="Arial", bold=True, size=10, color=C_HEAD)
        ws[f"A{row_idx}"].fill = fill(C_SECT)
        ws[f"A{row_idx}"].alignment = ca("left")
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1
        prev_module = module

    row_data = [sid, module, story, criteria, priority, platform, hrs,
                STATUS_DEFAULT, ASSIGNED_DEFAULT, ""]
    for ci, val in enumerate(row_data, start=1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.border = tb()
        is_even = (row_idx % 2 == 0)

        if ci == 1:  # ID
            c.font = bf(9, bold=True, color="1A1A1A")
            c.alignment = ca("center")
            c.fill = fill(C_ALT) if is_even else fill(C_WHITE)
        elif ci == 2:  # Module
            c.font = bf(9, bold=True)
            c.alignment = ca("left")
            c.fill = fill(C_ALT) if is_even else fill(C_WHITE)
        elif ci == 3:  # Story
            c.font = bf(9)
            c.alignment = ca("left")
            c.fill = fill(C_ALT) if is_even else fill(C_WHITE)
        elif ci == 4:  # Criteria
            c.font = Font(name="Arial", size=8.5, color="1A1A1A")
            c.alignment = ca("left")
            c.fill = fill(C_NOTE)
        elif ci == 5:  # Priority
            bg, fg = PMAP.get(priority, ("CCCCCC", "000000"))
            c.fill = fill(bg)
            c.font = Font(name="Arial", bold=True, size=8.5, color=fg)
            c.alignment = ca("center")
        elif ci == 6:  # Platform
            c.font = bf(9)
            c.alignment = ca("center")
            c.fill = fill(C_ALT) if is_even else fill(C_WHITE)
        elif ci == 7:  # Hours
            c.font = bf(9)
            c.alignment = ca("center")
            c.fill = fill(C_ALT) if is_even else fill(C_WHITE)
        elif ci == 8:  # Status
            c.font = bf(9)
            c.alignment = ca("center")
            c.fill = fill(C_DONE)
        else:
            c.font = bf(9)
            c.alignment = ca("left")
            c.fill = fill(C_ALT) if is_even else fill(C_WHITE)

    ws.row_dimensions[row_idx].height = 95
    row_idx += 1

# ════════════════════════════════════════════════════
# SHEET 3 — SPRINT PLAN
# ════════════════════════════════════════════════════
ws3 = wb.create_sheet("Sprint Plan")
ws3.sheet_view.showGridLines = False

mrow(ws3, 1, 7, "NUOVE — SUGGESTED SPRINT PLAN (8 WEEKS TO APP STORE)", C_NAV, sz=15)
ws3.row_dimensions[1].height = 40
mrow(ws3, 2, 7, "Each sprint = 1 week. Complete all P1 items before submitting to App Store / Play Store.", C_HEAD, sz=10)
ws3.row_dimensions[2].height = 22

sprints = [
    ("Sprint 1\nWeek 1",   "Infrastructure & Auth",
     "• INFRA-01: Upgrade Railway to paid\n• INFRA-02: Set up PostgreSQL\n• INFRA-03: Secure all env vars\n• INFRA-08: Email service (Resend + nuove.in DNS)\n• AUTH-01: Email/password signup + login\n• AUTH-04: JWT + refresh tokens\n• AUTH-05: Forgot password flow",
     "Backend live with DB. Users can sign up and log in.", 35),

    ("Sprint 2\nWeek 2",   "Core Features (Script + History)",
     "• AUTH-02: Google OAuth\n• AUTH-06: Onboarding flow\n• GEN-01 to GEN-06: Full script generator\n• HIST-01 to HIST-03: Script history (save/load/delete)\n• INFRA-05: UptimeRobot monitoring",
     "Users can generate, save and view scripts.", 38),

    ("Sprint 3\nWeek 3",   "Recorder + Crosspost",
     "• REC-01 to REC-06: Teleprompter & recorder (web)\n• REC-07: iOS codec fix\n• CROSS-01 to CROSS-04: Crosspost with captions\n• SEC-01 to SEC-06: Rate limiting + security headers",
     "Core features all working on web.", 40),

    ("Sprint 4\nWeek 4",   "Payments",
     "• PAY-01: Pricing page\n• PAY-02 to PAY-03: Razorpay live + webhook\n• PAY-04: Billing history\n• PAY-05: Cancel subscription\n• PAY-06: Free tier limits\n• PAY-07 to PAY-08: Payment emails\n• TEST-05: End-to-end payment QA",
     "Users can subscribe and pay. Revenue can be earned.", 44),

    ("Sprint 5\nWeek 5",   "Legal + Landing Page",
     "• LEGAL-01: Terms & Conditions live\n• LEGAL-02: Privacy Policy live\n• LEGAL-03: Consent checkbox on signup\n• LEGAL-04: AI content disclosure\n• LEGAL-05: GST/company registration (parallel, external)\n• PERF-06: Public landing page\n• PERF-07: Sitemap + robots.txt\n• PERF-01: Google Analytics 4",
     "App is legally compliant. Public landing page live.", 25),

    ("Sprint 6\nWeek 6",   "Mobile App Build",
     "• AUTH-03: Apple Sign In\n• AUTH-07: Splash screen\n• MOB-01: Capacitor iOS build\n• MOB-02: Capacitor Android build\n• MOB-03 to MOB-05: Mobile layout + permissions\n• MOB-08: Performance optimisation\n• MOB-10: App Store + Play Store assets",
     "Native apps built and running on real devices.", 42),

    ("Sprint 7\nWeek 7",   "QA + Bug Fixes",
     "• TEST-01: Manual QA checklist run\n• TEST-03: Real device testing (iPhone + Samsung)\n• TEST-04: Lighthouse scores\n• TEST-06: Cross-browser recording test\n• TEST-07: Network error handling\n• TEST-08: Accessibility audit\n• Fix all P1 bugs found",
     "All P1 and P2 bugs fixed. App ready for submission.", 36),

    ("Sprint 8\nWeek 8",   "App Store Submission",
     "• MOB-11: Final app icons + splash screens\n• MOB-10: Submit to Apple App Store (review: 1–3 days)\n• MOB-02: Submit to Google Play Store (review: 1–7 days)\n• NOTIF-01: Welcome email live\n• PROF-01 to PROF-06: Profile + settings complete\n• INFRA-04: CI/CD pipeline\n• Monitor: Sentry + Railway logs",
     "App submitted to both stores. 🚀", 30),
]

ws3.merge_cells("A3:G3")
ws3.row_dimensions[3].height = 8

hdrs3 = ["Sprint", "Theme", "Stories to Complete", "Definition of Done", "Est. Hours"]
for ci, h in enumerate(hdrs3, 1):
    c = ws3.cell(row=4, column=ci, value=h)
    c.font = hf(10)
    c.fill = fill(C_HEAD)
    c.alignment = ca()
    c.border = tb()
ws3.row_dimensions[4].height = 22

for ri, (sprint, theme, items, dod, hrs) in enumerate(sprints, start=5):
    row_data = [sprint, theme, items, dod, hrs]
    for ci, val in enumerate(row_data, start=1):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.border = tb()
        c.alignment = ca("left")
        c.font = bf(9, bold=(ci in [1,2,4]))
        if ri % 2 == 0:
            c.fill = fill(C_ALT)
        if ci == 1:
            c.fill = fill(C_HEAD)
            c.font = hf(10, color=C_WHITE)
            c.alignment = ca("center")
        if ci == 5:
            c.alignment = ca("center")
    ws3.row_dimensions[ri].height = 90

ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 55
ws3.column_dimensions["D"].width = 35
ws3.column_dimensions["E"].width = 12

# ── Save ──
out = "Nuove_PRD_UserStories.xlsx"
wb.save(out)
total = len(stories)
print(f"DONE. Saved: {out}")
print(f"Stories: {total}")
