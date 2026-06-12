from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

wb = Workbook()

# ── colour palette ────────────────────────────────────────────────
C_HEADER_BG   = "1A1A2E"   # deep navy
C_HEADER_FG   = "FFFFFF"
C_P1_BG       = "FF4C4C"   # red
C_P2_BG       = "FF9800"   # orange
C_P3_BG       = "4CAF50"   # green
C_P4_BG       = "2196F3"   # blue
C_P1_FG = C_P2_FG = C_P3_FG = C_P4_FG = "FFFFFF"
C_ALT_ROW     = "F0F4FF"   # light blue-tint for alternate rows
C_SECTION_BG  = "E8EAFF"   # section separator
C_WHITE       = "FFFFFF"
C_BORDER      = "C0C8E8"

PRIORITY_COLORS = {
    "P1 – Critical": (C_P1_BG, C_P1_FG),
    "P2 – High":     (C_P2_BG, C_P2_FG),
    "P3 – Medium":   (C_P3_BG, C_P3_FG),
    "P4 – Low":      (C_P4_BG, C_P4_FG),
}

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(sz=11, bold=True): return Font(name="Arial", bold=bold, size=sz, color=C_HEADER_FG)
def body_font(sz=10, bold=False): return Font(name="Arial", size=sz, bold=bold)
def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ══════════════════════════════════════════════════════════════════
# SHEET 1 – SUMMARY
# ══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📋 Summary"
ws1.sheet_view.showGridLines = False

# Title block
ws1.merge_cells("A1:H1")
ws1["A1"] = "Nuove – Production Readiness Tracker"
ws1["A1"].font = Font(name="Arial", bold=True, size=18, color="FFFFFF")
ws1["A1"].fill = fill(C_HEADER_BG)
ws1["A1"].alignment = center()
ws1.row_dimensions[1].height = 40

ws1.merge_cells("A2:H2")
ws1["A2"] = "Prepared for: Junior Developer Handoff  |  Prepared by: Senior Product Head  |  App: Nuove (AI Reel Coach)"
ws1["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
ws1["A2"].alignment = center()
ws1.row_dimensions[2].height = 20

ws1.merge_cells("A3:H3")  # spacer
ws1.row_dimensions[3].height = 8

# ── Legend ──
ws1.merge_cells("A4:H4")
ws1["A4"] = "PRIORITY LEGEND"
ws1["A4"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws1["A4"].fill = fill("3B3F72")
ws1["A4"].alignment = center()
ws1.row_dimensions[4].height = 22

legend = [
    ("P1 – Critical", C_P1_BG, "Must be done BEFORE launch. Blocking."),
    ("P2 – High",     C_P2_BG, "Do within first 2 weeks of launch."),
    ("P3 – Medium",   C_P3_BG, "Important – schedule within Month 1."),
    ("P4 – Low",      C_P4_BG, "Nice-to-have or post-Month-1."),
]
for i, (lbl, bg, desc) in enumerate(legend, start=5):
    ws1.merge_cells(f"A{i}:B{i}")
    ws1[f"A{i}"] = lbl
    ws1[f"A{i}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws1[f"A{i}"].fill = fill(bg)
    ws1[f"A{i}"].alignment = center()
    ws1.merge_cells(f"C{i}:H{i}")
    ws1[f"C{i}"] = desc
    ws1[f"C{i}"].font = body_font()
    ws1[f"C{i}"].alignment = left()
    ws1.row_dimensions[i].height = 20

# ── Category summary header ──
ws1.row_dimensions[9].height = 8  # spacer
ws1.merge_cells("A10:H10")
ws1["A10"] = "CATEGORY SUMMARY"
ws1["A10"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws1["A10"].fill = fill("3B3F72")
ws1["A10"].alignment = center()
ws1.row_dimensions[10].height = 22

headers_sum = ["Category", "Total Tasks", "P1", "P2", "P3", "P4", "Est. Hours", "Owner"]
for ci, h in enumerate(headers_sum, start=1):
    c = ws1.cell(row=11, column=ci, value=h)
    c.font = hdr_font(10)
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    c.border = thin_border()
ws1.row_dimensions[11].height = 22

categories = [
    ("1. Infrastructure & DevOps",   8,  3, 3, 2, 0,  24),
    ("2. Payments & Billing",        10, 4, 4, 2, 0,  40),
    ("3. Mobile Readiness",          14, 4, 5, 4, 1,  50),
    ("4. Security",                  10, 5, 3, 2, 0,  30),
    ("5. Feature Completion",        12, 3, 5, 3, 1,  55),
    ("6. Testing & QA",              11, 3, 4, 3, 1,  38),
    ("7. UI / UX Polish",             9, 2, 3, 3, 1,  28),
    ("8. Analytics & Monitoring",     7, 2, 3, 2, 0,  20),
    ("9. Email & Notifications",      6, 2, 2, 2, 0,  16),
    ("10. Performance",               7, 2, 3, 2, 0,  22),
    ("11. SEO & Social Meta",         5, 1, 2, 2, 0,  10),
    ("12. Legal & Compliance",        6, 3, 2, 1, 0,  18),
    ("13. Accessibility",             5, 1, 2, 2, 0,  12),
    ("14. PWA & Offline",             5, 1, 2, 1, 1,  15),
]

for ri, (cat, total, p1, p2, p3, p4, hrs) in enumerate(categories, start=12):
    row_data = [cat, total, p1, p2, p3, p4, hrs, "Junior Dev"]
    for ci, val in enumerate(row_data, start=1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.font = body_font(10, bold=(ci == 1))
        c.alignment = center() if ci != 1 else left()
        c.border = thin_border()
        if ri % 2 == 0:
            c.fill = fill(C_ALT_ROW)
        # colour code P columns
        if ci == 3 and p1: c.fill = fill("FFD5D5"); c.font = Font(name="Arial", size=10, bold=True, color=C_P1_BG)
        if ci == 4 and p2: c.fill = fill("FFE8CC"); c.font = Font(name="Arial", size=10, bold=True, color="E65C00")
        if ci == 5 and p3: c.fill = fill("DFFFDF"); c.font = Font(name="Arial", size=10, bold=True, color="2E7D32")
        if ci == 6 and p4: c.fill = fill("DCEFFF"); c.font = Font(name="Arial", size=10, bold=True, color="0D47A1")
    ws1.row_dimensions[ri].height = 20

# totals row
tr = 12 + len(categories)
total_tasks = sum(c[1] for c in categories)
total_p1    = sum(c[2] for c in categories)
total_p2    = sum(c[3] for c in categories)
total_p3    = sum(c[4] for c in categories)
total_p4    = sum(c[5] for c in categories)
total_hrs   = sum(c[6] for c in categories)
totals_row  = ["TOTAL", total_tasks, total_p1, total_p2, total_p3, total_p4, total_hrs, ""]
for ci, val in enumerate(totals_row, start=1):
    c = ws1.cell(row=tr, column=ci, value=val)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = fill(C_HEADER_BG)
    c.alignment = center() if ci != 1 else left()
    c.border = thin_border()
ws1.row_dimensions[tr].height = 22

# ── Infrastructure notes ──
note_row = tr + 2
ws1.merge_cells(f"A{note_row}:H{note_row}")
ws1[f"A{note_row}"] = "KEY INFRASTRUCTURE DECISIONS"
ws1[f"A{note_row}"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws1[f"A{note_row}"].fill = fill("3B3F72")
ws1[f"A{note_row}"].alignment = center()
ws1.row_dimensions[note_row].height = 22

infra_notes = [
    ("Frontend Hosting", "Vercel", "✅ Fine for production. Enable Analytics, set custom domain, configure rate-limit headers."),
    ("Backend Hosting",  "Railway", "⚠️ Free tier sleeps after 5 min inactivity – causes cold starts. Upgrade to Hobby ($5/mo) or Starter plan before launch."),
    ("Database",         "TBD",     "❗ Not confirmed. Add PostgreSQL or MongoDB (Railway add-on or Atlas free tier). Required for users, scripts, payments."),
    ("File Storage",     "TBD",     "❗ For recorded videos & uploads use Cloudflare R2 (free 10GB) or AWS S3."),
    ("CDN",              "Vercel Edge", "✅ Vercel handles CDN automatically for frontend assets."),
    ("Email",            "TBD",     "❗ Add Resend.com or SendGrid free tier for transactional emails (OTP, receipts)."),
    ("Payments",         "Razorpay","⚠️ Only test keys configured. Need live keys + webhook + GST details + T&C page before charging users."),
]
headers_inf = ["Component", "Current", "Recommendation / Action Required"]
for ci, h in enumerate(headers_inf, start=1):
    c = ws1.cell(row=note_row + 1, column=ci, value=h)
    c.font = hdr_font(10)
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    c.border = thin_border()
ws1.row_dimensions[note_row + 1].height = 20

for ri, (comp, cur, rec) in enumerate(infra_notes, start=note_row + 2):
    for ci, val in enumerate([comp, cur, rec], start=1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.font = body_font(10, bold=(ci == 1))
        c.alignment = left()
        c.border = thin_border()
        if ri % 2 == 0:
            c.fill = fill(C_ALT_ROW)
    ws1.row_dimensions[ri].height = 28

# Column widths – summary sheet
col_widths_s = [38, 12, 6, 6, 6, 6, 12, 16]
for i, w in enumerate(col_widths_s, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════
# SHEET 2 – ALL TASKS
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📝 All Tasks")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

# Title
ws2.merge_cells("A1:I1")
ws2["A1"] = "Nuove – Full Task List (115 Tasks)"
ws2["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
ws2["A1"].fill = fill(C_HEADER_BG)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 36

# Column headers
col_headers = ["#", "Category", "Task", "Details / Acceptance Criteria", "Priority", "Est. Hrs", "Status", "Assigned To", "Notes"]
for ci, h in enumerate(col_headers, start=1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.font = hdr_font(10)
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    c.border = thin_border()
ws2.row_dimensions[2].height = 22

# ── Task data ──
tasks = [
    # (Category, Task, Details, Priority, Est_Hrs)

    # 1. INFRASTRUCTURE & DEVOPS
    ("Infrastructure & DevOps", "Upgrade Railway to paid plan", "Free tier sleeps after 5 min. Upgrade to Hobby ($5/mo). Eliminates cold-start delays for real users.", "P1 – Critical", 1),
    ("Infrastructure & DevOps", "Add PostgreSQL database", "Railway PostgreSQL add-on OR Supabase free tier. Schema: users, scripts, sessions, payments, subscriptions.", "P1 – Critical", 3),
    ("Infrastructure & DevOps", "Configure custom domain on Vercel", "Point nuove.in (or .com) to Vercel deployment. Enable HTTPS auto-cert. Update CORS origins on backend.", "P1 – Critical", 2),
    ("Infrastructure & DevOps", "Set environment variables in prod", "Move all secrets (API keys, DB URLs, JWT secret, Razorpay keys) to Railway + Vercel env vars. No secrets in code.", "P2 – High", 2),
    ("Infrastructure & DevOps", "Set up CI/CD pipeline", "GitHub Actions: run lint + tests on PR, auto-deploy main to Railway+Vercel on merge. Fail fast on broken builds.", "P2 – High", 4),
    ("Infrastructure & DevOps", "Configure backend health-check endpoint", "GET /health returns 200 + uptime. Railway uses this to monitor. Prevents Railway from killing the dyno.", "P2 – High", 1),
    ("Infrastructure & DevOps", "Set up error tracking (Sentry)", "Add Sentry SDK to both frontend and backend. Configure source maps. Set up Slack/email alerts for P1 errors.", "P3 – Medium", 4),
    ("Infrastructure & DevOps", "Configure log aggregation", "Use Railway's built-in logs + Logtail (free tier). Tag logs with userId and route for debugging.", "P3 – Medium", 3),

    # 2. PAYMENTS & BILLING
    ("Payments & Billing", "Switch Razorpay to live keys", "Replace test key_id/key_secret with live credentials. Test end-to-end with ₹1 transaction before launch.", "P1 – Critical", 2),
    ("Payments & Billing", "Build subscription plans UI", "Show Free / Pro (₹X/mo) / Creator (₹X/mo) tiers with feature comparison table. Match brand design.", "P1 – Critical", 6),
    ("Payments & Billing", "Implement Razorpay webhook handler", "Verify webhook signature. Handle: payment.captured, subscription.activated, subscription.cancelled, payment.failed.", "P1 – Critical", 4),
    ("Payments & Billing", "Store subscription state in DB", "Users table: plan, subscription_id, subscription_status, next_billing_date. Gate features on plan.", "P1 – Critical", 3),
    ("Payments & Billing", "Build billing history page", "Show past invoices, next billing date, plan details. Allow cancel/upgrade from UI.", "P2 – High", 5),
    ("Payments & Billing", "Add GST invoice generation", "Indian law requires GST invoices. Integrate Razorpay's invoice API or generate PDF with company GSTIN.", "P2 – High", 4),
    ("Payments & Billing", "Implement free-tier usage limits", "Cap free users: e.g. 5 scripts/month, 3 recordings/month. Show upgrade prompt when limit hit.", "P2 – High", 4),
    ("Payments & Billing", "Add promo code / coupon support", "Razorpay supports discount codes. Build UI field + backend validation. For early-bird and influencer campaigns.", "P2 – High", 3),
    ("Payments & Billing", "Handle payment failure gracefully", "Show clear error message, don't lock user out immediately. Retry flow + email reminder.", "P3 – Medium", 3),
    ("Payments & Billing", "Refund & cancellation flow", "Admin panel or support email flow for refund requests. Razorpay refund API. Document policy in T&C.", "P3 – Medium", 5),

    # 3. MOBILE READINESS
    ("Mobile Readiness", "Audit all pages on 375px viewport", "Check every page on iPhone SE size. Fix overflows, unreadable text, broken layouts. Primary audience is mobile.", "P1 – Critical", 4),
    ("Mobile Readiness", "Fix camera/recorder on iOS Safari", "getUserMedia + MediaRecorder behaves differently on iOS. Test recording + playback. Handle codec differences (webm vs mp4).", "P1 – Critical", 6),
    ("Mobile Readiness", "Fix teleprompter scroll on touch devices", "Touch-scroll interferes with auto-scroll. Disable native scroll during recording. Test on Android Chrome + iOS Safari.", "P1 – Critical", 3),
    ("Mobile Readiness", "Test navigation on mobile (bottom nav)", "Bottom nav bar must be thumb-friendly (min 48px tap targets). Check active states, icons, labels.", "P1 – Critical", 2),
    ("Mobile Readiness", "Fix keyboard push-up layout on mobile", "When soft keyboard opens, it can break layouts. Use visual-viewport API to handle keyboard height.", "P2 – High", 4),
    ("Mobile Readiness", "Test Crosspost page on mobile", "Social platform buttons must be full-width and readable. Caption text area must resize correctly.", "P2 – High", 2),
    ("Mobile Readiness", "Test Generate page on mobile", "Dropdowns, tone selectors, script output card must all work on small screen.", "P2 – High", 2),
    ("Mobile Readiness", "Test Script History on mobile", "List items must be readable. Edit/delete actions accessible without hover.", "P2 – High", 2),
    ("Mobile Readiness", "Add mobile-specific touch gestures", "Swipe to dismiss modals. Pull-to-refresh on script history. Standard mobile UX patterns.", "P3 – Medium", 4),
    ("Mobile Readiness", "Optimize images for mobile", "All images served in WebP. Responsive srcset. No uncompressed PNGs. Reduces data usage for Indian mobile users.", "P3 – Medium", 3),
    ("Mobile Readiness", "Test on actual Android devices", "Test on low-end Android (2GB RAM, Chrome). Check performance, animations, font rendering.", "P3 – Medium", 3),
    ("Mobile Readiness", "Test on actual iOS devices", "Test on iPhone (Safari). Check recording, playback, all interactions.", "P3 – Medium", 3),
    ("Mobile Readiness", "Add haptic feedback (vibration API)", "Light vibration on record start/stop. Improves perceived quality on mobile.", "P4 – Low", 2),
    ("Mobile Readiness", "Consider React Native or Capacitor", "Long-term: a native app gets push notifications, better camera access. Evaluate after web is stable.", "P4 – Low", 0),

    # 4. SECURITY
    ("Security", "Implement JWT auth with refresh tokens", "Current auth may use short-lived tokens only. Add refresh token rotation. Store refresh token in httpOnly cookie.", "P1 – Critical", 5),
    ("Security", "Add rate limiting to all API routes", "Use express-rate-limit. Limits: /generate: 10/min, /auth: 5/min, /record: 20/min. Prevents abuse and AI API cost explosion.", "P1 – Critical", 3),
    ("Security", "Validate and sanitize all user inputs", "Use Zod or Joi for input validation on every backend route. Prevents injection attacks.", "P1 – Critical", 4),
    ("Security", "Add HTTPS-only headers", "Helmet.js: HSTS, X-Frame-Options, CSP, X-Content-Type-Options. 1-line setup protects from common attacks.", "P1 – Critical", 1),
    ("Security", "Protect AI API keys from front-end exposure", "Confirm no Anthropic/OpenAI keys are in frontend bundle. All AI calls must be server-side only.", "P1 – Critical", 2),
    ("Security", "Implement CORS whitelist", "Only allow requests from your Vercel domain + localhost. Currently may be open to all origins.", "P2 – High", 1),
    ("Security", "Add brute-force protection on login", "Lockout after 5 failed attempts. Implement with express-rate-limit or a per-user attempt counter in Redis/DB.", "P2 – High", 3),
    ("Security", "Encrypt sensitive data at rest", "User PII (email, name) should be stored with encryption or at minimum with a DB with encryption-at-rest enabled.", "P2 – High", 3),
    ("Security", "Security audit of dependencies", "Run `npm audit` on both frontend and backend. Fix all high/critical vulnerabilities. Add to CI/CD pipeline.", "P3 – Medium", 2),
    ("Security", "Add 2FA (optional but recommended)", "Optional TOTP (Google Authenticator) for Pro users. Adds trust signal for subscription tier.", "P3 – Medium", 6),

    # 5. FEATURE COMPLETION
    ("Feature Completion", "Complete user onboarding flow", "After signup: niche selection, language preference, sample script generation. First-run experience is critical for activation.", "P1 – Critical", 6),
    ("Feature Completion", "Build user settings page", "Change name, email, password, language preference, notification preferences. Currently missing or incomplete.", "P1 – Critical", 4),
    ("Feature Completion", "Complete script save/load flow", "Scripts saved to DB (not just localStorage). Associate scripts with user ID. Load on Scripts History page.", "P1 – Critical", 5),
    ("Feature Completion", "Integrate real social auth (Google/Apple)", "Currently may use email/password only. Google OAuth = fewer drop-offs. Apple required for iOS App Store.", "P2 – High", 5),
    ("Feature Completion", "Build admin dashboard", "View user count, active subscriptions, revenue, error logs. Minimum: read-only data view. Use Retool or build simple page.", "P2 – High", 8),
    ("Feature Completion", "Add notification system", "In-app notifications: script generated, subscription reminder, new feature alert. Badge count on bell icon.", "P2 – High", 5),
    ("Feature Completion", "Implement script sharing", "User can share script via link (read-only public URL). Good for viral growth.", "P2 – High", 4),
    ("Feature Completion", "Add multi-language caption generation", "Language selector in caption section works end-to-end. Supports at least: English, Hindi, Hinglish, Tamil, Telugu.", "P2 – High", 3),
    ("Feature Completion", "Complete Profile page", "Show script count, recording count, streak, joined date. Edit avatar/profile photo upload.", "P3 – Medium", 4),
    ("Feature Completion", "Build FAQ / Help page", "Simple static FAQ. Reduces support load. Answer top 10 questions: billing, recording, formats, etc.", "P3 – Medium", 3),
    ("Feature Completion", "Add feedback / bug report button", "Floating button on all pages. Submits to Airtable, Notion, or email. Free tools available.", "P3 – Medium", 2),
    ("Feature Completion", "Build Creator Dashboard widgets", "Weekly stats: scripts written, recordings made, streak. Show trend line. Motivates continued use.", "P4 – Low", 5),

    # 6. TESTING & QA
    ("Testing & QA", "Write unit tests for backend API routes", "Use Jest + Supertest. Test: /generate, /auth, /scripts, /payments. Min 70% coverage on critical paths.", "P1 – Critical", 8),
    ("Testing & QA", "Write integration tests for payment flow", "End-to-end test with Razorpay test mode: subscribe → webhook → feature unlock → cancel → feature lock.", "P1 – Critical", 5),
    ("Testing & QA", "Set up automated test runner in CI", "GitHub Actions runs all tests on every PR. Merge blocked if tests fail. Prevents regressions.", "P1 – Critical", 3),
    ("Testing & QA", "Manual QA test plan document", "Document step-by-step test cases for all features. Junior dev runs through before each release.", "P2 – High", 4),
    ("Testing & QA", "Cross-browser testing", "Test on: Chrome, Firefox, Safari (macOS), Safari (iOS), Chrome (Android). Note and fix differences.", "P2 – High", 4),
    ("Testing & QA", "Performance testing", "Use Lighthouse CI in GitHub Actions. Target: Performance >85, Accessibility >90, SEO >90 on mobile.", "P2 – High", 3),
    ("Testing & QA", "Load testing on backend", "Use k6 or Artillery. Simulate 100 concurrent users hitting /generate. Identify breaking point. Railway Hobby handles ~50 req/s.", "P3 – Medium", 4),
    ("Testing & QA", "Accessibility testing", "Run axe-core on all pages. Fix all critical and serious violations. WCAG 2.1 AA minimum.", "P3 – Medium", 4),
    ("Testing & QA", "User acceptance testing (UAT)", "Recruit 5 beta users (real content creators). Give them tasks. Record session with Hotjar. Fix top 3 pain points.", "P3 – Medium", 6),
    ("Testing & QA", "Error boundary testing", "Deliberately break API calls. Confirm error boundaries show friendly messages, not white screen of death.", "P4 – Low", 2),
    ("Testing & QA", "Visual regression testing", "Add Percy or Chromatic for visual diff on key pages. Prevents unintended UI regressions.", "P4 – Low", 4),

    # 7. UI / UX POLISH
    ("UI / UX Polish", "Add loading states to all async actions", "Every button that triggers an API call must show a spinner and be disabled while pending. No double-submit.", "P1 – Critical", 4),
    ("UI / UX Polish", "Add empty states to all list pages", "Script History empty: 'No scripts yet. Generate your first one!' with a CTA button. Same for other lists.", "P2 – High", 3),
    ("UI / UX Polish", "Toast / notification system", "Global toast for: ✓ Script saved, ✗ Error generating, ✓ Copied to clipboard. Use react-hot-toast.", "P2 – High", 2),
    ("UI / UX Polish", "Add skeleton loaders", "Replace blank screens during load with skeleton UI (pulsing grey blocks). Improves perceived performance.", "P2 – High", 3),
    ("UI / UX Polish", "Confirm destructive actions", "Delete script, cancel subscription: show modal 'Are you sure?' Prevents accidental data loss.", "P3 – Medium", 2),
    ("UI / UX Polish", "Add onboarding tooltips / coach marks", "First-time user sees highlighted tips on key features. Use react-joyride or shepherd.js.", "P3 – Medium", 4),
    ("UI / UX Polish", "Keyboard accessibility", "All interactive elements reachable by Tab. Enter/Space activates buttons. Esc closes modals.", "P3 – Medium", 4),
    ("UI / UX Polish", "Dark mode polish", "Audit all pages in dark mode. Fix any hard-coded light colours that look broken. Gradient text must be visible.", "P3 – Medium", 3),
    ("UI / UX Polish", "Micro-animation audit", "Review all page transitions and button interactions. Remove janky ones. Ensure consistent 200ms ease-out.", "P4 – Low", 3),

    # 8. ANALYTICS & MONITORING
    ("Analytics & Monitoring", "Add Vercel Analytics", "Enable Vercel's built-in Web Analytics. Free. Shows page views, top pages, countries.", "P1 – Critical", 1),
    ("Analytics & Monitoring", "Set up Google Analytics 4 or Posthog", "Track custom events: script_generated, recording_started, recording_completed, subscription_started.", "P1 – Critical", 3),
    ("Analytics & Monitoring", "Add uptime monitoring", "Use UptimeRobot (free). Ping /health every 5 min. Alert on Slack/email if down. Prevents Railway sleep on free tier.", "P2 – High", 1),
    ("Analytics & Monitoring", "Track funnel metrics", "Define funnel: signup → generate script → record → share. Measure drop-off at each step.", "P2 – High", 3),
    ("Analytics & Monitoring", "Add Hotjar or Microsoft Clarity", "Session recordings + heatmaps. Free tier is sufficient. Shows exactly where users are confused.", "P2 – High", 2),
    ("Analytics & Monitoring", "Build revenue dashboard", "Track MRR, churn rate, new subscriptions per day. Can use Razorpay dashboard + manual spreadsheet initially.", "P3 – Medium", 4),
    ("Analytics & Monitoring", "Set up alerting thresholds", "Alert if: error rate >1%, API latency >2s, new signups drop to 0 for 24h. Use Sentry + UptimeRobot.", "P3 – Medium", 3),

    # 9. EMAIL & NOTIFICATIONS
    ("Email & Notifications", "Set up transactional email (Resend or SendGrid)", "Choose provider. Add DNS records (SPF, DKIM). Prevents emails landing in spam.", "P1 – Critical", 2),
    ("Email & Notifications", "Welcome email on signup", "Triggered on new user creation. Include: app overview, top 3 features, link to first action.", "P1 – Critical", 2),
    ("Email & Notifications", "Payment receipt email", "Triggered on successful payment. Include: plan, amount, invoice link, support email.", "P2 – High", 2),
    ("Email & Notifications", "Subscription renewal reminder", "Email 3 days before renewal. Include: amount, cancel link, upgrade option.", "P2 – High", 2),
    ("Email & Notifications", "Weekly engagement email", "7-day summary: scripts written, streak. 'You wrote X scripts this week!' Drives retention.", "P3 – Medium", 4),
    ("Email & Notifications", "Password reset email", "Standard forgot-password flow. Secure token, 1-hour expiry. Test on all major email clients.", "P3 – Medium", 3),

    # 10. PERFORMANCE
    ("Performance", "Lazy-load route components", "Use React.lazy() + Suspense for all page components. Reduces initial bundle size. Critical for mobile 4G.", "P1 – Critical", 2),
    ("Performance", "Audit and reduce JavaScript bundle size", "Run `npx vite-bundle-analyzer`. Remove unused imports. Split large libraries (e.g. moment.js → dayjs).", "P1 – Critical", 4),
    ("Performance", "Optimise AI API response time", "Stream responses from AI backend instead of waiting for full response. Users see text appear progressively.", "P2 – High", 5),
    ("Performance", "Add Redis caching for repeated AI queries", "Cache results of identical prompts for 1 hour. Reduces AI API costs and speeds up repeat requests.", "P2 – High", 4),
    ("Performance", "Compress API responses (gzip/brotli)", "Enable compression middleware in Express. Reduces response payload size by ~70%.", "P3 – Medium", 1),
    ("Performance", "Preload critical fonts", "Add <link rel='preload'> for Dancing Script and any other fonts. Prevents layout shift on load.", "P3 – Medium", 1),
    ("Performance", "Image optimisation pipeline", "All uploaded images auto-compressed to WebP via Cloudflare or Sharp. Max 200KB per image.", "P3 – Medium", 3),

    # 11. SEO & SOCIAL META
    ("SEO & Social Meta", "Add meta tags to all public pages", "Title, description, og:image, og:title, og:description, twitter:card. Use react-helmet or Vite SSR.", "P1 – Critical", 3),
    ("SEO & Social Meta", "Create sitemap.xml and robots.txt", "Sitemap submitted to Google Search Console. robots.txt allows all public pages.", "P2 – High", 2),
    ("SEO & Social Meta", "Add structured data (JSON-LD)", "SoftwareApplication schema on homepage. Helps Google understand what the app does.", "P2 – High", 2),
    ("SEO & Social Meta", "Submit to Google Search Console", "Verify domain ownership. Submit sitemap. Monitor impressions and clicks from day 1.", "P3 – Medium", 1),
    ("SEO & Social Meta", "Create landing page / marketing page", "Public landing page (not behind auth): hero, features, pricing, testimonials. Currently all content is behind login.", "P3 – Medium", 8),

    # 12. LEGAL & COMPLIANCE
    ("Legal & Compliance", "Write Terms & Conditions page", "Must be live before accepting payments. Cover: usage, AI-generated content ownership, refund policy.", "P1 – Critical", 3),
    ("Legal & Compliance", "Write Privacy Policy page", "DPDP Act 2023 (India) compliance. State: what data collected, how stored, deletion rights.", "P1 – Critical", 3),
    ("Legal & Compliance", "Add cookie consent banner", "Required for EU users (GDPR). Use CookieYes free tier. Non-essential cookies only after consent.", "P1 – Critical", 1),
    ("Legal & Compliance", "Register company / GST", "Need GSTIN for Razorpay live account and invoices. Consult CA if not already registered.", "P2 – High", 0),
    ("Legal & Compliance", "AI content disclosure", "Add disclaimer: 'Scripts are AI-generated. Review before posting.' Required by emerging AI content laws.", "P2 – High", 1),
    ("Legal & Compliance", "Data deletion flow", "Users can request account + data deletion (DPDP requirement). Build 'Delete Account' in settings.", "P3 – Medium", 3),

    # 13. ACCESSIBILITY
    ("Accessibility", "Add alt text to all images and icons", "All <img> tags need descriptive alt. Icon-only buttons need aria-label. Screen reader usable.", "P1 – Critical", 2),
    ("Accessibility", "Fix colour contrast ratios", "Run WCAG contrast checker on all text/background combos. Minimum 4.5:1 for body text. Gradient text is risky.", "P2 – High", 3),
    ("Accessibility", "Add skip-to-content link", "Hidden link that appears on Tab press. Lets keyboard users skip nav to main content.", "P2 – High", 1),
    ("Accessibility", "ARIA roles on dynamic content", "Modal dialogs: role='dialog', aria-modal='true'. Live regions for AI streaming text: aria-live='polite'.", "P3 – Medium", 3),
    ("Accessibility", "Focus management in modals", "When modal opens, focus moves inside. When closed, focus returns to trigger. Prevents keyboard-user confusion.", "P3 – Medium", 2),

    # 14. PWA & OFFLINE
    ("PWA & Offline", "Add PWA manifest.json", "name, short_name, icons (192px + 512px), theme_color, background_color, display:'standalone'. Enables 'Add to Home Screen'.", "P1 – Critical", 2),
    ("PWA & Offline", "Register service worker", "Cache app shell (HTML, CSS, JS) for instant offline load. Use Workbox via Vite PWA plugin.", "P2 – High", 4),
    ("PWA & Offline", "Add offline fallback page", "When offline: show cached scripts and a friendly 'You're offline' message instead of browser error.", "P2 – High", 3),
    ("PWA & Offline", "Enable push notifications", "Ask permission after user generates first script. Send: 'Your weekly content tip is ready!' Drives re-engagement.", "P3 – Medium", 5),
    ("PWA & Offline", "Test install flow on Android + iOS", "Verify 'Add to Home Screen' prompt triggers on Android Chrome. iOS requires manual share > Add to Home.", "P4 – Low", 2),
]

STATUS_OPTIONS = "Not Started"
ASSIGNED = "Junior Dev"

for i, (cat, task, details, prio, hrs) in enumerate(tasks, start=1):
    row = i + 2  # offset for title + header
    row_data = [i, cat, task, details, prio, hrs, STATUS_OPTIONS, ASSIGNED, ""]
    for ci, val in enumerate(row_data, start=1):
        c = ws2.cell(row=row, column=ci, value=val)
        c.border = thin_border()
        c.alignment = left()
        if ci == 1:  # number
            c.alignment = center()
            c.font = body_font(9)
        elif ci == 2:  # category
            c.font = body_font(9, bold=True)
        elif ci == 3:  # task
            c.font = body_font(10, bold=True)
        elif ci == 4:  # details
            c.font = body_font(9)
        elif ci == 5:  # priority
            bg, fg = PRIORITY_COLORS.get(prio, (C_WHITE, "000000"))
            c.fill = fill(bg)
            c.font = Font(name="Arial", bold=True, size=9, color=fg)
            c.alignment = center()
        elif ci == 6:  # hours
            c.alignment = center()
            c.font = body_font(9)
        elif ci == 7:  # status
            c.fill = fill("FFF9E6")
            c.font = body_font(9)
        else:
            c.font = body_font(9)
        # alternate row shading (only on non-priority cells)
        if ci not in [5, 7] and i % 2 == 0:
            c.fill = fill(C_ALT_ROW)
    ws2.row_dimensions[row].height = 42

# add category section dividers (bold row when category changes)
prev_cat = None
for i, (cat, *_) in enumerate(tasks, start=1):
    if cat != prev_cat:
        row = i + 2
        for ci in range(1, 10):
            c = ws2.cell(row=row, column=ci)
            c.fill = fill(C_SECTION_BG)
            if ci == 2:
                c.font = Font(name="Arial", bold=True, size=10, color="1A1A2E")
        prev_cat = cat

# Column widths – task sheet
col_widths_t = [5, 22, 32, 55, 16, 8, 14, 14, 20]
for i, w in enumerate(col_widths_t, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ── Save ──
out_path = "Nuove_Production_Readiness_Tracker.xlsx"
wb.save(out_path)
print(f"DONE. Saved: {out_path}")
print(f"   Tasks: {len(tasks)} across {len(categories)} categories")
print(f"   Total estimated hours: {total_hrs}")
